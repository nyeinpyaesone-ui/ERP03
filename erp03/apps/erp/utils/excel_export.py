"""
Excel Export Service using openpyxl
Generates .xlsx files from query results
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import io

class ExcelExporter:
    """Generate Excel reports from data"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.cell_style = {'border': thin_border}
    
    def export(self, data: List[Dict[str, Any]], sheet_name: str = "Report") -> bytes:
        """Export list of dicts to Excel binary"""
        self.ws.title = sheet_name[:31]  # Excel limit
        
        if not data:
            return self._save()
        
        # Headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            
        # Data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                for key, style in self.cell_style.items():
                    setattr(cell, key, style)
        
        # Auto-width columns
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_num)
            for row in self.ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column].width = adjusted_width
            
        return self._save()
    
    def _save(self) -> bytes:
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
